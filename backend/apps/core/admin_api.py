from __future__ import annotations

import io
import json
import math
import os
import platform
import re
import shutil
import sqlite3
import subprocess
from calendar import monthrange
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any

from django.conf import settings
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.models import Group
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.http import require_GET, require_http_methods

from apps.audit.events import SUCCESSFUL_AUTH_EVENT_CODES
from apps.audit.models import OperationLog, UserActivityHour
from apps.audit.service import log_operation
from apps.catalog.models import (
    DATA_DOMAIN_TYPE_CHOICES,
    DataResource,
    DataResourceGroup,
    MapLayer,
    VectorDataset,
)
from apps.catalog.importer import ImportDataError, set_resource_access_groups
from apps.catalog.permissions import (
    filter_accessible,
    user_can_access,
    user_has_full_data_access,
)
from apps.core.api import (
    api_login_required,
    api_permission_required,
)
from apps.core.auth_views import serialize_user
from apps.core.emails import AccountEmailError, validate_account_email
from apps.core.config import (
    load_project_config,
    load_runtime_config_document,
    update_runtime_application_config,
)
from apps.core.initialization import (
    GUEST_GROUP_NAME,
    SUPERADMIN_GROUP_NAME,
    DEFAULT_USER_GROUP_NAME,
    ensure_guest_user,
    ensure_superadmin_defaults,
    is_builtin_group,
    is_guest_group,
    is_guest_user,
    is_initial_superadmin_user,
    is_platform_admin_group,
    is_superadmin_group,
    is_superadmin_user,
    protected_group_permissions,
    superadmin_group_locked_permissions,
)
from apps.core.backup_config import (
    BackupConfigError,
    serialize_backup_settings,
    update_backup_settings,
)
from apps.core.backup_service import (
    BackupServiceError,
    backup_scope_summaries,
    local_backup_download_path,
    serialize_backup_run,
    start_backup_run,
    test_backup_target,
)
from apps.core.models import BackupRun, RoleApplication, SystemSetting, UserProfile
from apps.core.passwords import generate_password, password_validation_errors
from apps.core.permissions import (
    FEATURE_PERMISSION_NAMES,
    FEATURE_PERMISSIONS,
    disabled_feature_permissions,
    direct_feature_permissions,
    effective_feature_permissions,
    feature_denied_response,
    feature_permission_queryset,
    granted_feature_permissions,
    has_feature_perm,
)
from apps.core.principal_visibility import (
    group_is_visible_to,
    selectable_access_groups_for,
    user_is_visible_to,
    visible_group_ids_for,
    visible_groups_for,
    visible_operation_logs_for,
    visible_users_for,
)
from apps.core.role_applications import (
    RoleApplicationReviewError,
    review_role_application,
    serialize_role_application,
)
from apps.core.storage import (
    StoragePathError,
    app_path,
    research_path,
    table_data_path,
    validate_vector_layer_name,
    vector_geopackage_path,
    vector_original_path,
)
from apps.raster.models import RasterDataset


@require_GET
@api_login_required
def admin_profile(request):
    return JsonResponse(_serialize_profile(request.user))


@require_http_methods(["POST"])
@api_login_required
def update_admin_profile(request):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    user = request.user
    profile = _ensure_profile(user)
    # 用户名在创建时确定，不允许修改
    display_name = payload.get("displayName")
    email = payload.get("email")
    avatar_url = payload.get("avatarUrl")
    department = payload.get("department")

    if display_name is not None:
        user.first_name = str(display_name).strip()
        user.last_name = ""
    if email is not None:
        if is_guest_user(user):
            return JsonResponse(
                {"detail": f"{GUEST_GROUP_NAME}账号不能修改邮箱"}, status=400
            )
        try:
            user.email = validate_account_email(email, exclude_user_id=user.id)
        except AccountEmailError as exc:
            return JsonResponse({"detail": str(exc)}, status=400)
        profile.normalized_email = user.email
    if avatar_url is not None:
        profile.avatar_url = str(avatar_url).strip()
    if department is not None:
        profile.department = str(department).strip()

    try:
        with transaction.atomic():
            user.save()
            profile.save()
    except IntegrityError:
        log_operation(
            request.user,
            "用户设置",
            "更新个人资料",
            "failed",
            "个人资料保存失败",
            request,
        )
        return JsonResponse({"detail": "保存失败"}, status=400)

    log_operation(
        request.user,
        "用户设置",
        "更新个人资料",
        "success",
        "个人资料已更新",
        request,
    )
    return JsonResponse(_serialize_profile(user))


@require_http_methods(["POST"])
@api_login_required
def upload_avatar(request):
    """上传用户头像"""
    if "avatar" not in request.FILES:
        return JsonResponse({"detail": "请选择头像文件"}, status=400)

    avatar_file = request.FILES["avatar"]

    # 验证文件格式
    allowed_types = ["image/jpeg", "image/png"]
    if avatar_file.content_type not in allowed_types:
        return JsonResponse({"detail": "头像格式仅支持 JPG 和 PNG"}, status=400)

    # 验证文件大小 (2MB)
    max_size = 2 * 1024 * 1024
    if avatar_file.size > max_size:
        return JsonResponse({"detail": "头像文件大小不能超过 2MB"}, status=400)

    # 读取文件数据
    file_data = avatar_file.read()

    # 压缩图片至合适尺寸
    from PIL import Image, UnidentifiedImageError

    try:
        image = Image.open(io.BytesIO(file_data))
        # 转换为RGB模式（如果是RGBA）
        if image.mode in ("RGBA", "LA"):
            background = Image.new("RGB", image.size, (255, 255, 255))
            background.paste(image, mask=image.split()[-1])
            image = background
        elif image.mode != "RGB":
            image = image.convert("RGB")

        # 压缩到合适尺寸 (最大 300x300)
        max_size_pixels = (300, 300)
        image.thumbnail(max_size_pixels, Image.Resampling.LANCZOS)

        # 保存为JPEG格式
        output = io.BytesIO()
        image.save(output, format="JPEG", quality=85)
        file_data = output.getvalue()
        content_type = "image/jpeg"
    except (UnidentifiedImageError, OSError):
        return JsonResponse({"detail": "头像文件不是有效图片"}, status=400)

    # 保存到数据库
    user = request.user
    profile = _ensure_profile(user)
    profile.avatar_data = file_data
    profile.avatar_content_type = content_type
    # 清除URL头像
    profile.avatar_url = ""
    profile.save(
        update_fields=["avatar_data", "avatar_content_type", "avatar_url", "updated_at"]
    )

    log_operation(
        request.user,
        "用户设置",
        "上传头像",
        "success",
        "上传头像成功",
        request,
    )

    return JsonResponse(_serialize_profile(user))


@require_http_methods(["GET"])
@api_login_required
def get_avatar(request, user_id: int):
    """获取用户头像"""
    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if not user_is_visible_to(request.user, user):
        return JsonResponse({"detail": "用户不存在"}, status=404)

    try:
        profile = user.profile
        if profile.avatar_data:
            from django.http import HttpResponse

            response = HttpResponse(
                profile.avatar_data, content_type=profile.avatar_content_type
            )
            response["Content-Disposition"] = f'inline; filename="avatar_{user_id}.jpg"'
            return response
    except ObjectDoesNotExist:
        pass

    return JsonResponse({"detail": "用户未设置头像"}, status=404)


@require_http_methods(["POST"])
@api_login_required
def update_admin_profile_permissions(request):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    disabled = payload.get("disabledPermissions")
    if not isinstance(disabled, list):
        return JsonResponse({"detail": "disabledPermissions 必须是数组"}, status=400)
    disabled_set = {str(permission) for permission in disabled}
    granted = granted_feature_permissions(request.user)
    invalid = sorted(disabled_set - granted)
    if invalid:
        return JsonResponse(
            {"detail": f"不能关闭未授予的权限：{', '.join(invalid)}"},
            status=400,
        )
    locked_permissions = superadmin_group_locked_permissions()
    if is_superadmin_user(request.user) and disabled_set & locked_permissions:
        return JsonResponse(
            {"detail": f"{SUPERADMIN_GROUP_NAME}不能关闭系统锁定权限"},
            status=400,
        )

    profile = _ensure_profile(request.user)
    profile.disabled_permissions = sorted(disabled_set)
    profile.save(update_fields=["disabled_permissions", "updated_at"])
    log_operation(
        request.user,
        "用户设置",
        "更新个人权限开关",
        "success",
        f"关闭 {len(disabled_set)} 项权限",
        request,
    )
    return JsonResponse(_serialize_profile(request.user))


@require_http_methods(["POST"])
@api_login_required
def update_admin_profile_password(request):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    current_password = str(payload.get("currentPassword", ""))
    new_password = str(payload.get("newPassword", ""))
    password_confirm = str(payload.get("passwordConfirm", ""))
    if not request.user.check_password(current_password):
        log_operation(
            request.user,
            "认证授权",
            "修改密码",
            "failed",
            "当前密码验证失败",
            request,
        )
        return JsonResponse({"detail": "当前密码不正确"}, status=400)
    if new_password != password_confirm:
        log_operation(
            request.user,
            "认证授权",
            "修改密码",
            "failed",
            "两次输入的新密码不一致",
            request,
        )
        return JsonResponse({"detail": "两次输入的新密码不一致"}, status=400)
    if current_password == new_password:
        log_operation(
            request.user,
            "认证授权",
            "修改密码",
            "failed",
            "新密码与当前密码相同",
            request,
        )
        return JsonResponse({"detail": "新密码不能与当前密码相同"}, status=400)
    password_errors = password_validation_errors(new_password, user=request.user)
    if password_errors:
        log_operation(
            request.user,
            "认证授权",
            "修改密码",
            "failed",
            "新密码强度校验失败",
            request,
        )
        return JsonResponse({"detail": "；".join(password_errors)}, status=400)

    request.user.set_password(new_password)
    request.user.save(update_fields=["password"])
    update_session_auth_hash(request, request.user)
    log_operation(
        request.user,
        "认证授权",
        "修改密码",
        "success",
        "用户已修改密码",
        request,
    )
    return JsonResponse({"detail": "密码已更新"})


@require_http_methods(["GET", "POST"])
@api_permission_required("core.manage_auth")
def group_list(request):
    if request.method == "GET":
        ensure_superadmin_defaults(
            create_account=False, attach_existing_superusers=False
        )
        ensure_guest_user()
        return JsonResponse(
            {
                "items": [
                    _serialize_group(group, request.user)
                    for group in visible_groups_for(_groups(), request.user)
                ],
                "availablePermissions": _available_permissions(),
            }
        )
    if not has_feature_perm(request.user, "core.manage_feature_permissions"):
        return JsonResponse({"detail": "当前用户无权限配置角色"}, status=403)

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    name = _required_string(payload.get("name"), "name")
    if isinstance(name, JsonResponse):
        return name
    permissions = _permission_names(payload.get("permissions", []))
    if isinstance(permissions, JsonResponse):
        return permissions

    try:
        group = Group.objects.create(name=name)
        _set_group_feature_permissions(group, permissions)
    except IntegrityError:
        return JsonResponse({"detail": "角色名称已存在"}, status=400)

    log_operation(request.user, "认证授权", "创建角色", "success", group.name, request)
    return JsonResponse(_serialize_group(group, request.user), status=201)


@require_http_methods(["POST"])
@api_permission_required("core.manage_auth")
def group_detail(request, group_id: int):
    try:
        group = Group.objects.get(pk=group_id)
    except Group.DoesNotExist:
        return JsonResponse({"detail": "角色不存在"}, status=404)
    if not group_is_visible_to(request.user, group):
        return JsonResponse({"detail": "角色不存在"}, status=404)

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    # 检查是否是删除操作
    if payload.get("action") == "delete":
        if is_builtin_group(group):
            return JsonResponse({"detail": "系统内置角色不能删除"}, status=400)
        if group.user_set.exists():
            return JsonResponse({"detail": "角色仍有关联用户，不能删除"}, status=400)
        group_name = group.name
        group.delete()
        log_operation(
            request.user, "认证授权", "删除角色", "success", group_name, request
        )
        return JsonResponse({"detail": "角色已删除"})

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    if "name" in payload:
        name = _required_string(payload.get("name"), "name")
        if isinstance(name, JsonResponse):
            return name
        if is_builtin_group(group) and name != group.name:
            return JsonResponse(
                {"detail": f"{group.name}角色名称不能修改"},
                status=400,
            )
        group.name = name
    if "permissions" in payload:
        permissions = _permission_names(payload.get("permissions"))
        if isinstance(permissions, JsonResponse):
            return permissions
        if is_superadmin_group(group):
            locked = superadmin_group_locked_permissions()
            if locked - set(permissions):
                return JsonResponse(
                    {"detail": f"{SUPERADMIN_GROUP_NAME}角色必须保留系统锁定权限"},
                    status=400,
                )
            permissions = protected_group_permissions()
        if is_builtin_group(group) and not is_superadmin_group(group):
            permissions = sorted(set(permissions))
        _set_group_feature_permissions(group, permissions)
    try:
        group.save()
    except IntegrityError:
        return JsonResponse({"detail": "角色名称已存在"}, status=400)
    log_operation(request.user, "认证授权", "更新角色", "success", group.name, request)
    return JsonResponse(_serialize_group(group, request.user))


@require_http_methods(["GET", "POST"])
@api_permission_required("core.manage_auth")
def user_list(request):
    User = get_user_model()
    if request.method == "POST":
        if not has_feature_perm(request.user, "core.create_user"):
            return JsonResponse({"detail": "当前用户无新建用户权限"}, status=403)
        payload = _json_payload(request)
        if isinstance(payload, JsonResponse):
            return payload
        result = _create_admin_user(User, payload, request.user)
        if isinstance(result, JsonResponse):
            return result
        created, generated_password = result
        log_operation(
            request.user,
            "认证授权",
            "创建用户",
            "success",
            created.get_username(),
            request,
        )
        response_data = _serialize_admin_user(created, request.user)
        response_data["generatedPassword"] = generated_password
        return JsonResponse(response_data, status=201)

    users = visible_users_for(
        User.objects.prefetch_related(
            "groups",
            "user_permissions__content_type",
            "groups__permissions__content_type",
            "profile",
        ).order_by("id"),
        request.user,
    )
    return JsonResponse(
        {"items": [_serialize_admin_user(user, request.user) for user in users]}
    )


@require_GET
@api_permission_required("core.manage_auth")
def role_application_list(request):
    status = str(request.GET.get("status", "")).strip()
    valid_statuses = set(RoleApplication.Status.values)
    if status and status not in valid_statuses:
        return JsonResponse({"detail": "角色申请状态不存在"}, status=400)

    User = get_user_model()
    visible_user_ids = visible_users_for(User.objects.all(), request.user).values("id")
    applications = RoleApplication.objects.filter(user_id__in=visible_user_ids)
    if status:
        applications = applications.filter(status=status)
    applications = applications.select_related(
        "user", "user__profile", "reviewer", "reviewer__profile"
    )
    return JsonResponse(
        {
            "items": [
                serialize_role_application(application, include_user=True)
                for application in applications
            ]
        }
    )


@require_http_methods(["POST"])
@api_permission_required("core.manage_auth")
def role_application_review(request, application_id: int):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    try:
        application = RoleApplication.objects.select_related("user").get(
            pk=application_id
        )
    except RoleApplication.DoesNotExist:
        return JsonResponse({"detail": "角色申请不存在"}, status=404)
    if not user_is_visible_to(request.user, application.user):
        return JsonResponse({"detail": "角色申请不存在"}, status=404)

    action = str(payload.get("action", "")).strip()
    review_note = str(payload.get("reviewNote", "")).strip()
    if len(review_note) > 500:
        return JsonResponse({"detail": "审核说明不能超过 500 个字符"}, status=400)
    try:
        application = review_role_application(
            application_id,
            reviewer=request.user,
            action=action,
            review_note=review_note,
        )
    except RoleApplication.DoesNotExist:
        return JsonResponse({"detail": "角色申请不存在"}, status=404)
    except RoleApplicationReviewError as exc:
        return JsonResponse({"detail": str(exc)}, status=400)

    action_label = "通过" if action == "approve" else "拒绝"
    log_operation(
        request.user,
        "认证授权",
        "审核角色申请",
        "success",
        f"{action_label} {application.user.get_username()} 的科研用户申请",
        request,
    )
    return JsonResponse(serialize_role_application(application, include_user=True))


@require_http_methods(["POST"])
@api_permission_required("core.manage_auth")
def user_detail(request, user_id: int):
    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if not user_is_visible_to(request.user, user):
        return JsonResponse({"detail": "用户不存在"}, status=404)

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    # 检查是否是删除操作
    if payload.get("action") == "delete":
        if user.pk == request.user.pk:
            return JsonResponse({"detail": "不能删除当前登录用户"}, status=400)
        if is_initial_superadmin_user(user):
            return JsonResponse({"detail": "初始化管理员不能删除"}, status=400)
        if is_guest_user(user):
            return JsonResponse(
                {"detail": f"{GUEST_GROUP_NAME}账号不能删除"}, status=400
            )
        username = user.get_username()
        user.delete()
        log_operation(
            request.user,
            "认证授权",
            "删除用户",
            "success",
            username,
            request,
        )
        return JsonResponse({"detail": "用户已删除"})

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    if "isActive" not in payload:
        return JsonResponse({"detail": "缺少 isActive"}, status=400)
    is_active = bool(payload["isActive"])
    if not is_active and user.pk == request.user.pk:
        return JsonResponse({"detail": "不能停用当前登录用户"}, status=400)
    if not is_active and is_initial_superadmin_user(user):
        return JsonResponse({"detail": "初始化管理员不能停用"}, status=400)
    if not is_active and is_guest_user(user):
        return JsonResponse({"detail": f"{GUEST_GROUP_NAME}账号不能停用"}, status=400)
    user.is_active = is_active
    user.save(update_fields=["is_active"])
    log_operation(
        request.user,
        "认证授权",
        "更新用户状态",
        "success",
        f"{user.get_username()} {'启用' if is_active else '停用'}",
        request,
    )
    return JsonResponse(_serialize_admin_user(user, request.user))


@require_http_methods(["POST"])
@api_permission_required("core.manage_auth")
def reset_user_password(request, user_id: int):
    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if not user_is_visible_to(request.user, user):
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if user.pk == request.user.pk:
        return JsonResponse({"detail": "不能重置当前登录用户密码"}, status=400)
    if is_guest_user(user):
        return JsonResponse(
            {"detail": f"{GUEST_GROUP_NAME}账号不能重置密码"}, status=400
        )

    password = generate_password()
    user.set_password(password)
    user.save(update_fields=["password"])
    log_operation(
        request.user,
        "认证授权",
        "重置用户密码",
        "success",
        user.get_username(),
        request,
    )
    response_data = _serialize_admin_user(user, request.user)
    response_data["generatedPassword"] = password
    return JsonResponse(response_data)


@require_http_methods(["POST"])
@api_permission_required("core.manage_auth")
def update_user_groups(request, user_id: int):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    group_ids = payload.get("groupIds")
    if not isinstance(group_ids, list):
        return JsonResponse({"detail": "groupIds 必须是数组"}, status=400)

    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if not user_is_visible_to(request.user, user):
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if user.pk == request.user.pk:
        return JsonResponse({"detail": "不能修改当前登录用户的角色"}, status=400)
    if is_superadmin_user(user):
        return JsonResponse(
            {"detail": f"不能修改{SUPERADMIN_GROUP_NAME}的角色"}, status=400
        )
    if is_guest_user(user):
        return JsonResponse(
            {"detail": f"{GUEST_GROUP_NAME}账号不能修改角色"}, status=400
        )

    try:
        normalized_group_ids = {int(group_id) for group_id in group_ids}
    except (TypeError, ValueError):
        return JsonResponse({"detail": "groupIds 必须是整数数组"}, status=400)
    if is_initial_superadmin_user(user):
        _, protected_group = ensure_superadmin_defaults(
            create_account=False, attach_existing_superusers=False
        )
        normalized_group_ids.add(protected_group.id)
    elif not normalized_group_ids:
        return JsonResponse({"detail": "角色为必选项"}, status=400)

    groups = list(
        visible_groups_for(
            Group.objects.filter(id__in=normalized_group_ids), request.user
        )
    )
    if len(groups) != len(normalized_group_ids):
        return JsonResponse({"detail": "包含不存在的角色"}, status=400)
    if not is_initial_superadmin_user(user) and any(
        is_superadmin_group(group) for group in groups
    ):
        return JsonResponse(
            {
                "detail": f"不能将{DEFAULT_USER_GROUP_NAME}加入{SUPERADMIN_GROUP_NAME}角色"
            },
            status=400,
        )
    if not is_superadmin_user(request.user) and any(
        is_platform_admin_group(group) for group in groups
    ):
        return JsonResponse(
            {"detail": "只有超级管理员可以分配平台管理员角色"}, status=400
        )
    user.groups.set(groups)
    log_operation(
        request.user,
        "认证授权",
        "设置角色",
        "success",
        user.get_username(),
        request,
    )
    return JsonResponse(_serialize_admin_user(user, request.user))


@require_http_methods(["POST"])
@api_permission_required("core.manage_auth")
def update_user_permissions(request, user_id: int):
    if not has_feature_perm(request.user, "core.manage_feature_permissions"):
        return JsonResponse({"detail": "当前用户无权限配置用户权限"}, status=403)

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    if "directPermissions" not in payload:
        return JsonResponse({"detail": "directPermissions 必须是数组"}, status=400)
    permissions = _permission_names(payload.get("directPermissions"))
    if isinstance(permissions, JsonResponse):
        return permissions
    operation_log_group_ids = _operation_log_group_ids_payload(
        payload.get("operationLogGroupIds", []), request.user
    )
    if isinstance(operation_log_group_ids, JsonResponse):
        return operation_log_group_ids
    disabled_permissions = _permission_names(payload.get("disabledPermissions", []))
    if isinstance(disabled_permissions, JsonResponse):
        return disabled_permissions

    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if not user_is_visible_to(request.user, user):
        return JsonResponse({"detail": "用户不存在"}, status=404)
    if user.pk == request.user.pk:
        return JsonResponse({"detail": "请到用户设置中修改自己的权限"}, status=400)
    if is_guest_user(user):
        return JsonResponse(
            {"detail": f"{GUEST_GROUP_NAME}账号不能设置直授权限"}, status=400
        )

    _set_user_feature_permissions(user, permissions)
    _set_disabled_permissions(user, disabled_permissions)
    _set_operation_log_group_ids(user, operation_log_group_ids)
    log_operation(
        request.user,
        "认证授权",
        "设置用户权限",
        "success",
        user.get_username(),
        request,
    )
    return JsonResponse(_serialize_admin_user(user, request.user))


@require_http_methods(["GET", "POST"])
@api_permission_required("core.manage_system_settings")
def admin_settings(request):
    if request.method == "GET":
        return JsonResponse(_serialize_application_settings(request.user))

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload

    patch: dict[str, Any] = {}
    if "systemName" in payload:
        system_name = _required_string(payload["systemName"], "systemName")
        if isinstance(system_name, JsonResponse):
            return system_name
        patch.setdefault("system", {})["name"] = system_name
    if "allowRegistration" in payload:
        allow_registration = _boolean_value(
            payload["allowRegistration"], "allowRegistration"
        )
        if isinstance(allow_registration, JsonResponse):
            return allow_registration
        patch.setdefault("system", {})["allow_registration"] = allow_registration
    if "map" in payload:
        map_patch = _map_patch(payload["map"])
        if isinstance(map_patch, JsonResponse):
            return map_patch
        patch["map"] = map_patch
    if "limits" in payload:
        limits_patch = _limits_patch(payload["limits"])
        if isinstance(limits_patch, JsonResponse):
            return limits_patch
        patch["limits"] = limits_patch
    if "raster" in payload:
        raster_patch = _raster_patch(payload["raster"])
        if isinstance(raster_patch, JsonResponse):
            return raster_patch
        patch["raster"] = raster_patch

    if patch:
        update_runtime_application_config(settings.PROJECT_CONFIG, patch)
        _reload_runtime_project_config()
        if "system" in patch and "allow_registration" in patch["system"]:
            SystemSetting.objects.update_or_create(
                pk=1,
                defaults={"allow_registration": patch["system"]["allow_registration"]},
            )
        log_operation(request.user, "系统设置", "保存配置", "success", "", request)

    return JsonResponse(_serialize_application_settings(request.user))


def _reload_runtime_project_config() -> None:
    config = load_project_config(
        settings.PROJECT_CONFIG.config_path,
        program_root=settings.PROGRAM_ROOT,
    )
    settings.PROJECT_CONFIG = config
    settings.DATA_UPLOAD_MAX_MEMORY_SIZE = None


@require_GET
def admin_backup_overview(request):
    denied = _backup_superadmin_denied(request)
    if denied:
        return denied
    active_runs = BackupRun.objects.select_related("created_by").filter(
        status__in=[BackupRun.Status.QUEUED, BackupRun.Status.RUNNING]
    )[:5]
    recent_runs = BackupRun.objects.select_related("created_by").order_by(
        "-created_at"
    )[:8]
    return JsonResponse(
        {
            "settings": serialize_backup_settings(),
            "summaries": backup_scope_summaries(),
            "activeRuns": [serialize_backup_run(run) for run in active_runs],
            "recentRuns": [serialize_backup_run(run) for run in recent_runs],
            "generatedAt": timezone.localtime().isoformat(),
        }
    )


@require_http_methods(["GET", "POST"])
def admin_backup_settings(request):
    denied = _backup_superadmin_denied(request)
    if denied:
        return denied
    if request.method == "GET":
        return JsonResponse(serialize_backup_settings())
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    try:
        updated = update_backup_settings(payload)
    except BackupConfigError as exc:
        log_operation(
            request.user,
            "数据备份",
            "保存备份配置",
            "failed",
            str(exc),
            request,
        )
        return JsonResponse({"detail": str(exc)}, status=400)
    _reload_runtime_project_config()
    log_operation(request.user, "数据备份", "保存备份配置", "success", "", request)
    return JsonResponse(serialize_backup_settings(updated))


@require_http_methods(["POST"])
def admin_backup_target_test(request):
    denied = _backup_superadmin_denied(request)
    if denied:
        return denied
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    target_type = str(payload.get("targetType") or "").strip()
    try:
        test_backup_target(
            target_type,
            local=payload.get("local"),
            object_storage=payload.get("objectStorage"),
        )
    except (BackupConfigError, BackupServiceError, Exception) as exc:
        log_operation(
            request.user,
            "数据备份",
            "测试备份目标",
            "failed",
            str(exc),
            request,
        )
        return JsonResponse(
            {
                "targetType": target_type or "object_storage",
                "status": "failed",
                "message": str(exc),
                "testedAt": timezone.localtime().isoformat(),
            }
        )
    log_operation(
        request.user,
        "数据备份",
        "测试备份目标",
        "success",
        target_type,
        request,
    )
    return JsonResponse(
        {
            "targetType": target_type,
            "status": "success",
            "message": "备份目标连接测试成功",
            "testedAt": timezone.localtime().isoformat(),
        }
    )


@require_http_methods(["GET", "POST"])
def admin_backup_runs(request):
    denied = _backup_superadmin_denied(request)
    if denied:
        return denied
    if request.method == "GET":
        queryset = BackupRun.objects.select_related("created_by").order_by(
            "-created_at"
        )
        queryset = _filter_backup_runs(queryset, request.GET)
        if isinstance(queryset, JsonResponse):
            return queryset
        total = queryset.count()
        current = _positive_query_int(request.GET.get("current"), default=1)
        page_size = _positive_query_int(request.GET.get("pageSize"), default=20)
        if isinstance(current, JsonResponse):
            return current
        if isinstance(page_size, JsonResponse):
            return page_size
        start = (current - 1) * page_size
        end = start + page_size
        return JsonResponse(
            {
                "items": [serialize_backup_run(run) for run in queryset[start:end]],
                "total": total,
            }
        )

    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    try:
        run = start_backup_run(
            plan_type=str(payload.get("planType") or "").strip(),
            target_type=str(payload.get("targetType") or "").strip() or None,
            trigger=BackupRun.Trigger.MANUAL,
            user=request.user,
            include_logs=payload.get("includeLogs"),
            run_async=True,
        )
    except BackupServiceError as exc:
        log_operation(
            request.user,
            "数据备份",
            "发起备份任务",
            "failed",
            str(exc),
            request,
        )
        return JsonResponse({"detail": str(exc)}, status=400)
    log_operation(
        request.user,
        "数据备份",
        "发起备份任务",
        "success",
        f"任务 {run.id}：{run.plan_type} -> {run.target_type}",
        request,
    )
    return JsonResponse(serialize_backup_run(run), status=202)


@require_GET
def admin_backup_run_detail(request, run_id: int):
    denied = _backup_superadmin_denied(request)
    if denied:
        return denied
    try:
        run = BackupRun.objects.select_related("created_by").get(pk=run_id)
    except BackupRun.DoesNotExist:
        return JsonResponse({"detail": "备份任务不存在"}, status=404)
    return JsonResponse(serialize_backup_run(run))


@require_GET
def admin_backup_run_download(request, run_id: int):
    denied = _backup_superadmin_denied(request)
    if denied:
        return denied
    path = local_backup_download_path(run_id)
    if isinstance(path, JsonResponse):
        return path
    return FileResponse(
        path.open("rb"),
        as_attachment=True,
        filename=path.name,
        content_type="application/zip",
    )


@require_GET
@api_login_required
def admin_data_resources(request):
    if not (
        has_feature_perm(request.user, "catalog.view_dataresource")
        or has_feature_perm(request.user, "catalog.change_dataresource")
        or has_feature_perm(request.user, "catalog.delete_dataresource")
        or has_feature_perm(request.user, "catalog.export_dataresource")
        or has_feature_perm(request.user, "catalog.add_dataresource")
    ):
        return feature_denied_response(request.user)
    queryset = _filter_admin_data_resources(
        DataResource.objects.select_related("category", "maintainer")
        .select_related("inventory_group")
        .prefetch_related("access_groups", "map_layers")
        .order_by("-updated_at", "name"),
        request.GET,
    )
    if (
        has_feature_perm(request.user, "catalog.view_dataresource")
        or has_feature_perm(request.user, "catalog.change_dataresource")
        or has_feature_perm(request.user, "catalog.delete_dataresource")
        or has_feature_perm(request.user, "catalog.export_dataresource")
    ):
        queryset = filter_accessible(queryset, request.user)
    else:
        queryset = queryset.filter(maintainer=request.user)
    summary_queryset = DataResource.objects.filter(
        pk__in=queryset.order_by().values("pk")
    )
    summary = _admin_data_resource_summary(summary_queryset)
    inventory_groups = list(DataResourceGroup.objects.order_by("sort_order", "id"))
    total = summary["total"]
    current = _positive_query_int(request.GET.get("current"), default=1)
    page_size = _positive_query_int(request.GET.get("pageSize"), default=20)
    if isinstance(current, JsonResponse):
        return current
    if isinstance(page_size, JsonResponse):
        return page_size
    start = (current - 1) * page_size
    end = start + page_size
    return JsonResponse(
        {
            "items": [
                _serialize_admin_data_resource(resource, request.user)
                for resource in queryset[start:end]
            ],
            "total": total,
            "summary": summary,
            "groupSummaries": _admin_data_resource_group_summaries(
                summary_queryset, inventory_groups, summary
            ),
            "availableAccessGroups": [
                _serialize_access_group(group)
                for group in selectable_access_groups_for(
                    Group.objects.order_by("name"), request.user
                )
            ],
            "inventoryGroups": [
                _serialize_data_resource_group(group) for group in inventory_groups
            ],
        }
    )


@require_http_methods(["POST"])
@api_permission_required("catalog.change_dataresource")
def admin_data_resource_groups(request):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    name = _clean_resource_group_name(payload.get("name"))
    if isinstance(name, JsonResponse):
        return name
    try:
        group = DataResourceGroup.objects.create(name=name)
    except IntegrityError:
        return JsonResponse({"detail": "组别名称已存在"}, status=400)
    log_operation(
        request.user,
        "数据管理",
        "新增数据组别",
        "success",
        group.name,
        request,
    )
    return JsonResponse(_serialize_data_resource_group(group), status=201)


@require_http_methods(["POST"])
@api_permission_required("catalog.change_dataresource")
def admin_data_resource_group_detail(request, group_id: int):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    try:
        group = DataResourceGroup.objects.get(pk=group_id)
    except DataResourceGroup.DoesNotExist:
        return JsonResponse({"detail": "数据组别不存在"}, status=404)
    action = str(payload.get("action", "update")).strip()
    if action == "delete":
        name = group.name
        affected_count = group.resources.count()
        group.delete()
        log_operation(
            request.user,
            "数据管理",
            "删除数据组别",
            "success",
            f"{name}；{affected_count} 项数据进入默认分组",
            request,
        )
        return JsonResponse({"detail": "数据组别已删除"})
    if action != "update":
        return JsonResponse({"detail": "不支持的数据组别操作"}, status=400)
    name = _clean_resource_group_name(payload.get("name"))
    if isinstance(name, JsonResponse):
        return name
    group.name = name
    try:
        group.save(update_fields=["name", "updated_at"])
    except IntegrityError:
        return JsonResponse({"detail": "组别名称已存在"}, status=400)
    log_operation(
        request.user,
        "数据管理",
        "更新数据组别",
        "success",
        group.name,
        request,
    )
    return JsonResponse(_serialize_data_resource_group(group))


@require_http_methods(["POST"])
@api_login_required
def admin_data_resource_detail(request, resource_id: int):
    payload = _json_payload(request)
    if isinstance(payload, JsonResponse):
        return payload
    try:
        resource = (
            DataResource.objects.select_related("category", "maintainer")
            .select_related("inventory_group")
            .prefetch_related("access_groups", "map_layers")
            .get(pk=resource_id)
        )
    except DataResource.DoesNotExist:
        return JsonResponse({"detail": "数据资源不存在"}, status=404)
    if not user_can_access(resource, request.user):
        return JsonResponse({"detail": "数据资源不存在"}, status=404)

    action = str(payload.get("action", "update")).strip()
    can_change = has_feature_perm(request.user, "catalog.change_dataresource")
    can_delete = has_feature_perm(request.user, "catalog.delete_dataresource")
    can_update_access = can_change or resource.maintainer_id == request.user.id
    if action == "delete":
        if not can_delete:
            return feature_denied_response(request.user)
        return _delete_admin_data_resource(request, resource, payload)
    if action not in {
        "update",
        "setStatus",
        "saveVisualization",
        "updateAccess",
        "updateInventoryGroup",
    }:
        return JsonResponse({"detail": "不支持的数据资源操作"}, status=400)
    if action == "updateAccess":
        if not can_update_access:
            return feature_denied_response(request.user)
    elif not can_change:
        return feature_denied_response(request.user)
    if "accessGroupIds" in payload and not (can_change or can_update_access):
        return feature_denied_response(request.user)

    update_fields = []
    changed_access = False
    changed_inventory_group = False
    if action in {"update", "setStatus"} and "status" in payload:
        status = str(payload.get("status", "")).strip()
        if status not in DataResource.Status.values:
            return JsonResponse(
                {"detail": "status 仅支持 active 或 inactive"}, status=400
            )
        resource.status = status
        update_fields.append("status")
        MapLayer.objects.filter(data_resource=resource).update(
            is_active=status == DataResource.Status.ACTIVE
        )

    if action in {"update", "updateAccess"} and "accessGroupIds" in payload:
        group_ids = _normalize_group_ids(payload.get("accessGroupIds"))
        if isinstance(group_ids, JsonResponse):
            return group_ids
        try:
            set_resource_access_groups(resource, group_ids, viewer=request.user)
        except ImportDataError as exc:
            return JsonResponse({"detail": str(exc)}, status=400)
        changed_access = True

    if action in {"update", "updateInventoryGroup"} and "inventoryGroupId" in payload:
        inventory_group_id = payload.get("inventoryGroupId")
        if inventory_group_id is None:
            resource.inventory_group = None
        else:
            try:
                resource.inventory_group = DataResourceGroup.objects.get(
                    pk=int(inventory_group_id)
                )
            except (TypeError, ValueError, DataResourceGroup.DoesNotExist):
                return JsonResponse({"detail": "数据组别不存在"}, status=400)
        update_fields.append("inventory_group")
        changed_inventory_group = True

    if action in {"update", "saveVisualization"} and "visualization" in payload:
        visualization = payload.get("visualization")
        if not isinstance(visualization, dict):
            return JsonResponse(
                {"detail": "visualization 必须是 JSON 对象"}, status=400
            )
        resource.default_visualization = visualization
        update_fields.append("default_visualization")
        layer_error = _sync_default_visualization_layer(resource, visualization)
        if isinstance(layer_error, JsonResponse):
            return layer_error

    if update_fields:
        update_fields.append("updated_at")
        resource.save(update_fields=sorted(set(update_fields)))

    if not update_fields and not changed_access and not changed_inventory_group:
        return JsonResponse({"detail": "没有可更新的数据资源字段"}, status=400)

    log_operation(
        request.user,
        "数据管理",
        _admin_resource_action_label(action, payload),
        "success",
        resource.name,
        request,
        target_type="data_resource",
        target_id=resource.id,
        target_code=resource.code,
        target_name=resource.name,
    )
    resource.refresh_from_db()
    updated_resource = (
        DataResource.objects.select_related("category", "maintainer")
        .select_related("inventory_group")
        .prefetch_related("access_groups", "map_layers")
        .get(pk=resource.id)
    )
    return JsonResponse(_serialize_admin_data_resource(updated_resource, request.user))


@require_GET
@api_permission_required("catalog.export_dataresource")
def admin_data_resources_export(request):
    queryset = _filter_admin_data_resources(
        DataResource.objects.select_related("category", "maintainer")
        .select_related("inventory_group")
        .prefetch_related("access_groups", "map_layers")
        .order_by("-updated_at", "name"),
        request.GET,
    )
    queryset = filter_accessible(queryset, request.user)
    resources = list(queryset)
    export_format = str(request.GET.get("format", "csv")).strip().lower()
    if export_format == "xlsx":
        response = _data_resources_xlsx_response(resources, request.user)
    elif export_format == "csv":
        response = _data_resources_csv_response(resources, request.user)
    else:
        return JsonResponse({"detail": "format 仅支持 csv 或 xlsx"}, status=400)

    log_operation(
        request.user,
        "数据管理",
        "导出存量数据",
        "success",
        f"{export_format.upper()} {len(resources)} 条",
        request,
    )
    return response


@require_GET
@api_login_required
def admin_operation_logs(request):
    logs = OperationLog.objects.select_related("user").order_by("-created_at")
    logs = _scope_operation_logs(logs, request.user)
    logs = _filter_operation_logs(logs, request.GET)
    total = logs.count()
    current = _positive_query_int(request.GET.get("current"), default=1)
    page_size = _positive_query_int(request.GET.get("pageSize"), default=20)
    if isinstance(current, JsonResponse):
        return current
    if isinstance(page_size, JsonResponse):
        return page_size
    start = (current - 1) * page_size
    end = start + page_size
    return JsonResponse(
        {
            "items": [_serialize_operation_log(log) for log in logs[start:end]],
            "total": total,
        }
    )


@require_GET
@api_permission_required("core.view_system_logs")
def admin_system_logs(request):
    lines = _positive_query_int(request.GET.get("lines"), default=500)
    if isinstance(lines, JsonResponse):
        return lines
    lines = min(lines, 2000)

    files = _available_system_log_files()
    requested_file = str(request.GET.get("file", "")).strip()
    selected_file = requested_file or (files[0]["name"] if files else "")
    if selected_file and not any(item["name"] == selected_file for item in files):
        return JsonResponse({"detail": "日志文件不存在"}, status=404)

    return JsonResponse(
        {
            "files": files,
            "selectedFile": selected_file,
            "lines": lines,
            "content": _tail_system_log_file(selected_file, lines)
            if selected_file
            else "",
            "generatedAt": timezone.localtime().isoformat(),
        }
    )


@require_GET
@api_login_required
def admin_dashboard(request):
    period = _active_period(request.GET.get("period"))
    if isinstance(period, JsonResponse):
        return period
    cards: dict[str, Any] = {}
    if has_feature_perm(request.user, "core.view_dashboard_resource_card"):
        cards["resources"] = {
            "total": DataResource.objects.count(),
            "active": DataResource.objects.filter(
                status=DataResource.Status.ACTIVE
            ).count(),
        }
    if has_feature_perm(request.user, "core.view_dashboard_layer_card"):
        cards["layers"] = {
            "total": MapLayer.objects.count(),
            "active": MapLayer.objects.filter(is_active=True).count(),
        }
    if has_feature_perm(request.user, "core.view_dashboard_raster_card"):
        cards["rasters"] = {
            "resources": DataResource.objects.filter(
                data_type=DataResource.DataType.RASTER
            ).count(),
            "datasets": RasterDataset.objects.count(),
            "layers": MapLayer.objects.filter(
                layer_type=MapLayer.LayerType.RASTER
            ).count(),
        }
    if has_feature_perm(request.user, "core.view_dashboard_user_card"):
        User = get_user_model()
        users = visible_users_for(User.objects.all(), request.user)
        groups = visible_groups_for(Group.objects.all(), request.user)
        cards["users"] = {
            "total": users.count(),
            "active": users.filter(is_active=True).count(),
            "disabled": users.filter(is_active=False).count(),
            "groups": groups.count(),
            "vectorResources": DataResource.objects.filter(
                data_type=DataResource.DataType.VECTOR
            ).count(),
            "tableResources": DataResource.objects.filter(
                data_type=DataResource.DataType.TABLE
            ).count(),
        }
    if has_feature_perm(request.user, "core.view_dashboard_active_users_card"):
        period_start, period_end = _active_period_bounds(period)
        User = get_user_model()
        visible_user_ids = visible_users_for(User.objects.all(), request.user).values(
            "id"
        )
        activity_hours = UserActivityHour.objects.filter(
            user_id__in=visible_user_ids,
            bucket_start__gte=period_start,
            bucket_start__lt=period_end,
        )
        login_logs = visible_operation_logs_for(
            OperationLog.objects.filter(
                event_code__in=SUCCESSFUL_AUTH_EVENT_CODES,
                status=OperationLog.Status.SUCCESS,
                created_at__gte=period_start,
                created_at__lt=period_end,
                user_id__isnull=False,
            ),
            request.user,
        )
        cards["activeUsers"] = {
            "period": period,
            "rangeStart": timezone.localtime(period_start).date().isoformat(),
            "rangeEnd": (timezone.localtime(period_end) - timedelta(days=1))
            .date()
            .isoformat(),
            "count": activity_hours.values("user_id").distinct().count(),
            "loginCount": login_logs.count(),
            "series": _active_user_series(activity_hours, period, period_start),
            "ranking": _active_user_ranking(login_logs),
        }
    can_view_data_overview = has_feature_perm(request.user, "core.view_data_overview")
    cards["dataOverview"] = _data_overview_card(
        request.user,
        include_visible=can_view_data_overview,
        include_uploaders=can_view_data_overview
        and user_has_full_data_access(request.user),
    )

    return JsonResponse(
        {
            "generatedAt": timezone.localtime().isoformat(),
            "cards": cards,
        }
    )


@require_GET
@api_permission_required("core.view_dashboard_system_card")
def admin_dashboard_server(request):
    return JsonResponse(_server_snapshot(request.user))


def _active_period(value: Any) -> str | JsonResponse:
    period = str(value or "day").strip()
    if period not in {"day", "week", "month"}:
        return JsonResponse({"detail": "period 仅支持 day、week、month"}, status=400)
    return period


def _active_period_bounds(period: str):
    today = timezone.localdate()
    if period == "week":
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=7)
    elif period == "month":
        start_date = today.replace(day=1)
        end_date = start_date + timedelta(days=monthrange(today.year, today.month)[1])
    else:
        start_date = today
        end_date = today + timedelta(days=1)
    start = timezone.make_aware(datetime.combine(start_date, time.min))
    end = timezone.make_aware(datetime.combine(end_date, time.min))
    return start, end


def _active_user_series(
    activity_hours, period: str, period_start
) -> list[dict[str, Any]]:
    if period == "day":
        counts = {hour: set() for hour in range(24)}
        for user_id, bucket_start in activity_hours.values_list(
            "user_id", "bucket_start"
        ):
            counts[timezone.localtime(bucket_start).hour].add(user_id)
        return [
            {"key": str(hour), "label": f"{hour:02d}:00", "count": len(user_ids)}
            for hour, user_ids in counts.items()
        ]

    start_date = timezone.localtime(period_start).date()
    days = 7 if period == "week" else monthrange(start_date.year, start_date.month)[1]
    counts = {start_date + timedelta(days=offset): set() for offset in range(days)}
    for user_id, bucket_start in activity_hours.values_list("user_id", "bucket_start"):
        date = timezone.localtime(bucket_start).date()
        if date in counts:
            counts[date].add(user_id)
    return [
        {
            "key": date.isoformat(),
            "label": date.strftime("%m-%d"),
            "count": len(user_ids),
        }
        for date, user_ids in counts.items()
    ]


def _active_user_ranking(login_logs) -> list[dict[str, Any]]:
    ranked_logs = (
        login_logs.values("user_id", "user__username", "user__first_name")
        .annotate(login_count=Count("id"))
        .order_by("-login_count", "user__username")[:5]
    )
    return [
        {
            "userId": row["user_id"],
            "displayName": row["user__first_name"] or row["user__username"],
            "username": row["user__username"],
            "loginCount": row["login_count"],
        }
        for row in ranked_logs
    ]


def _data_overview_card(
    user, *, include_visible: bool, include_uploaders: bool
) -> dict[str, Any]:
    own_uploads = DataResource.objects.filter(maintainer=user)
    card: dict[str, Any] = {
        "ownUploads": _data_overview_scope(own_uploads),
    }
    if include_visible:
        all_resources = DataResource.objects.all()
        visible_resources = filter_accessible(DataResource.objects.all(), user)
        card.update(
            {
                **_data_overview_scope(all_resources, include_spatial=False),
                "visibleResources": _data_overview_scope(visible_resources),
            }
        )
    if include_uploaders:
        card["uploaders"] = _uploader_stats()
    return card


MAX_DASHBOARD_SPATIAL_RESOURCES = 80
DASHBOARD_HEATMAP_COLUMNS = 8
DASHBOARD_HEATMAP_ROWS = 5


def _data_overview_scope(queryset, *, include_spatial: bool = True) -> dict[str, Any]:
    aggregate = queryset.aggregate(
        total_size=Sum("size_bytes"),
        total_items=Sum("item_count"),
    )
    scope = {
        "totalResources": queryset.count(),
        "activeResources": queryset.filter(status=DataResource.Status.ACTIVE).count(),
        "totalSizeBytes": int(aggregate["total_size"] or 0),
        "totalItemCount": int(aggregate["total_items"] or 0),
        "typeBreakdown": _data_type_breakdown(queryset),
    }
    if include_spatial:
        scope["spatialSummary"] = _data_spatial_summary(queryset)
    return scope


def _data_type_breakdown(queryset) -> list[dict[str, Any]]:
    rows = (
        queryset.values("data_type")
        .annotate(
            count=Count("id"),
            size_bytes=Sum("size_bytes"),
            item_count=Sum("item_count"),
        )
        .order_by("data_type")
    )
    return [
        {
            "dataType": row["data_type"],
            "count": row["count"],
            "sizeBytes": int(row["size_bytes"] or 0),
            "itemCount": int(row["item_count"] or 0),
        }
        for row in rows
    ]


def _data_spatial_summary(queryset) -> dict[str, Any]:
    resources = queryset.select_related("maintainer").prefetch_related(
        "map_layers", "raster_datasets"
    )
    extents = []
    missing_count = 0
    for resource in resources:
        bounds = _resource_bounds(resource)
        if bounds is None:
            missing_count += 1
            continue
        min_x, min_y, max_x, max_y = bounds
        center = [(min_x + max_x) / 2, (min_y + max_y) / 2]
        coverage_area = _bounds_area_km2(bounds)
        extents.append(
            {
                "resourceId": resource.id,
                "name": resource.name,
                "dataType": resource.data_type,
                "bounds": [round(value, 6) for value in bounds],
                "center": [round(value, 6) for value in center],
                "coverageAreaKm2": round(coverage_area, 3),
                "sizeBytes": resource.size_bytes,
                "itemCount": resource.item_count,
                "coordinateSystem": resource.coordinate_system,
                "uploaderName": _user_display_name(resource.maintainer),
                "updatedAt": timezone.localtime(resource.updated_at).isoformat(),
            }
        )

    total_bounds = _union_bounds(item["bounds"] for item in extents)
    ranking = sorted(
        extents,
        key=lambda item: (-item["coverageAreaKm2"], item["name"], item["resourceId"]),
    )
    displayed_extents = ranking[:MAX_DASHBOARD_SPATIAL_RESOURCES]
    return {
        "spatialResourceCount": len(extents),
        "missingSpatialResourceCount": missing_count,
        "totalBounds": total_bounds or [],
        "resourceExtents": displayed_extents,
        "resourceExtentsTruncated": len(extents) > len(displayed_extents),
        "coverageRanking": ranking[:10],
        "heatmapCells": _spatial_heatmap_cells(extents, total_bounds),
    }


def _resource_bounds(resource: DataResource) -> list[float] | None:
    bounds = _normalize_bounds(resource.spatial_extent)
    if bounds is not None:
        return bounds
    for layer in resource.map_layers.all():
        bounds = _normalize_bounds(layer.bounds)
        if bounds is not None:
            return bounds
    for dataset in resource.raster_datasets.all():
        bounds = _normalize_bounds(dataset.bounds_4326)
        if bounds is not None:
            return bounds
    return None


def _normalize_bounds(value: Any) -> list[float] | None:
    if isinstance(value, str):
        parts = re.split(r"[,，\s]+", value.strip())
        raw_values = [part for part in parts if part]
    elif isinstance(value, list | tuple):
        raw_values = list(value)
    else:
        return None
    if len(raw_values) != 4:
        return None
    try:
        min_x, min_y, max_x, max_y = [float(item) for item in raw_values]
    except (TypeError, ValueError):
        return None
    values = [min_x, min_y, max_x, max_y]
    if not all(math.isfinite(item) for item in values):
        return None
    if min_x > max_x:
        min_x, max_x = max_x, min_x
    if min_y > max_y:
        min_y, max_y = max_y, min_y
    if min_x == max_x or min_y == max_y:
        return None
    if min_x < -180 or max_x > 180 or min_y < -90 or max_y > 90:
        return None
    return [min_x, min_y, max_x, max_y]


def _union_bounds(bounds_items) -> list[float] | None:
    bounds_list = list(bounds_items)
    if not bounds_list:
        return None
    return [
        round(min(bounds[0] for bounds in bounds_list), 6),
        round(min(bounds[1] for bounds in bounds_list), 6),
        round(max(bounds[2] for bounds in bounds_list), 6),
        round(max(bounds[3] for bounds in bounds_list), 6),
    ]


def _bounds_area_km2(bounds: list[float]) -> float:
    min_x, min_y, max_x, max_y = bounds
    center_latitude = (min_y + max_y) / 2
    lat_km = max_y - min_y
    lon_km = (max_x - min_x) * max(math.cos(math.radians(center_latitude)), 0.01)
    return abs(lat_km * lon_km) * 111.32 * 111.32


def _spatial_heatmap_cells(
    extents: list[dict[str, Any]], total_bounds: list[float] | None
) -> list[dict[str, Any]]:
    if not extents or not total_bounds:
        return []
    min_x, min_y, max_x, max_y = total_bounds
    span_x = max(max_x - min_x, 0.000001)
    span_y = max(max_y - min_y, 0.000001)
    cells: dict[tuple[int, int], dict[str, Any]] = {}
    for item in extents:
        center_x, center_y = item["center"]
        column = min(
            DASHBOARD_HEATMAP_COLUMNS - 1,
            max(0, int(((center_x - min_x) / span_x) * DASHBOARD_HEATMAP_COLUMNS)),
        )
        row = min(
            DASHBOARD_HEATMAP_ROWS - 1,
            max(0, int(((center_y - min_y) / span_y) * DASHBOARD_HEATMAP_ROWS)),
        )
        cell = cells.setdefault(
            (column, row),
            {
                "column": column,
                "row": row,
                "bounds": [
                    min_x + (span_x / DASHBOARD_HEATMAP_COLUMNS) * column,
                    min_y + (span_y / DASHBOARD_HEATMAP_ROWS) * row,
                    min_x + (span_x / DASHBOARD_HEATMAP_COLUMNS) * (column + 1),
                    min_y + (span_y / DASHBOARD_HEATMAP_ROWS) * (row + 1),
                ],
                "resourceCount": 0,
                "itemCount": 0,
            },
        )
        cell["resourceCount"] += 1
        cell["itemCount"] += item["itemCount"]
    return [
        {
            **cell,
            "bounds": [round(value, 6) for value in cell["bounds"]],
        }
        for cell in sorted(
            cells.values(),
            key=lambda item: (-item["resourceCount"], item["column"], item["row"]),
        )
    ]


def _user_display_name(user) -> str:
    if not user:
        return "未记录"
    return user.get_full_name() or user.get_username() or "未记录"


def _uploader_stats() -> list[dict[str, Any]]:
    rows = (
        DataResource.objects.values(
            "maintainer_id", "maintainer__username", "maintainer__first_name"
        )
        .annotate(
            resource_count=Count("id"),
            size_bytes=Sum("size_bytes"),
            item_count=Sum("item_count"),
        )
        .order_by("-resource_count", "maintainer__username")
    )
    return [
        {
            "user": {
                "id": row["maintainer_id"],
                "username": row["maintainer__username"] or "",
                "displayName": row["maintainer__first_name"]
                or row["maintainer__username"]
                or "未记录",
            },
            "resourceCount": row["resource_count"],
            "sizeBytes": int(row["size_bytes"] or 0),
            "itemCount": int(row["item_count"] or 0),
        }
        for row in rows
    ]


def _server_snapshot(user) -> dict[str, Any]:
    cards: dict[str, Any] = {}
    if has_feature_perm(user, "core.view_dashboard_system_card"):
        cards["cpu"] = _cpu_snapshot()
        cards["memory"] = _memory_snapshot()
        cards["disks"] = _disk_snapshot()
    return {
        "generatedAt": timezone.localtime().isoformat(),
        "hostname": platform.node(),
        "platform": platform.platform(),
        "cards": cards,
    }


def _filter_admin_data_resources(queryset, params):
    query = str(params.get("q", "")).strip()
    data_type = str(params.get("dataType", "")).strip()
    status = str(params.get("status", "")).strip()
    category = str(params.get("category", "")).strip()
    source = str(params.get("source", "")).strip()
    provider = str(params.get("provider", "")).strip()
    date_from = parse_date(str(params.get("dateFrom", "")).strip())
    date_to = parse_date(str(params.get("dateTo", "")).strip())

    if query:
        queryset = queryset.filter(
            Q(name__icontains=query)
            | Q(code__icontains=query)
            | Q(description__icontains=query)
            | Q(source__icontains=query)
            | Q(provider__icontains=query)
        )
    if data_type:
        queryset = queryset.filter(data_type=data_type)
    if status:
        queryset = queryset.filter(status=status)
    if category:
        queryset = queryset.filter(category__code=category)
    if source:
        queryset = queryset.filter(source__icontains=source)
    if provider:
        queryset = queryset.filter(provider__icontains=provider)
    if date_from:
        queryset = queryset.filter(data_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(data_date__lte=date_to)
    return queryset


def _admin_data_resource_summary(queryset) -> dict[str, int]:
    stats = _admin_data_resource_stat_values(queryset)
    non_superadmin_groups = Group.objects.exclude(name=SUPERADMIN_GROUP_NAME).values(
        "id"
    )
    restricted_count = (
        queryset.filter(access_groups__in=non_superadmin_groups).distinct().count()
    )
    return {
        "total": stats["resourceCount"],
        "activeCount": stats["activeCount"],
        "inactiveCount": stats["inactiveCount"],
        "restrictedCount": restricted_count,
        "sizeBytes": stats["sizeBytes"],
        "itemCount": stats["itemCount"],
    }


def _admin_data_resource_group_summaries(
    queryset,
    inventory_groups: list[DataResourceGroup],
    summary: dict[str, int],
) -> list[dict[str, Any]]:
    all_stats = {
        "resourceCount": summary["total"],
        "activeCount": summary["activeCount"],
        "inactiveCount": summary["inactiveCount"],
        "sizeBytes": summary["sizeBytes"],
        "itemCount": summary["itemCount"],
    }
    domain_codes = [code for code, _label in DATA_DOMAIN_TYPE_CHOICES]
    domain_stats = {code: _empty_admin_data_resource_stats() for code in domain_codes}
    for row in (
        queryset.order_by()
        .values("domain_type")
        .annotate(**_admin_data_resource_stat_annotations())
    ):
        domain_type = row.pop("domain_type") or "other"
        if domain_type not in domain_stats:
            domain_type = "other"
        _merge_admin_data_resource_stats(domain_stats[domain_type], row)

    custom_stats = {
        row["inventory_group_id"]: row
        for row in (
            queryset.filter(inventory_group_id__isnull=False)
            .order_by()
            .values("inventory_group_id")
            .annotate(**_admin_data_resource_stat_annotations())
        )
    }
    summaries = [
        _admin_data_resource_group_summary(key="__all__", kind="all", stats=all_stats)
    ]
    summaries.extend(
        _admin_data_resource_group_summary(
            key=f"__domain__:{domain_type}",
            kind="business",
            domain_type=domain_type,
            stats=domain_stats[domain_type],
        )
        for domain_type in domain_codes
    )
    summaries.extend(
        _admin_data_resource_group_summary(
            key=f"__custom__:{group.id}",
            kind="custom",
            inventory_group_id=group.id,
            stats=custom_stats.get(group.id, _empty_admin_data_resource_stats()),
        )
        for group in inventory_groups
    )
    return summaries


def _admin_data_resource_stat_annotations() -> dict[str, Any]:
    return {
        "resourceCount": Count("id"),
        "activeCount": Count("id", filter=Q(status=DataResource.Status.ACTIVE)),
        "inactiveCount": Count("id", filter=Q(status=DataResource.Status.INACTIVE)),
        "sizeBytes": Sum("size_bytes"),
        "itemCount": Sum("item_count"),
    }


def _admin_data_resource_stat_values(queryset) -> dict[str, int]:
    values = queryset.aggregate(**_admin_data_resource_stat_annotations())
    return {key: int(value or 0) for key, value in values.items()}


def _empty_admin_data_resource_stats() -> dict[str, int]:
    return {
        "resourceCount": 0,
        "activeCount": 0,
        "inactiveCount": 0,
        "sizeBytes": 0,
        "itemCount": 0,
    }


def _merge_admin_data_resource_stats(
    target: dict[str, int], source: dict[str, Any]
) -> None:
    for key in target:
        target[key] += int(source.get(key) or 0)


def _admin_data_resource_group_summary(
    *,
    key: str,
    kind: str,
    stats: dict[str, Any],
    domain_type: str | None = None,
    inventory_group_id: int | None = None,
) -> dict[str, Any]:
    return {
        "key": key,
        "kind": kind,
        "domainType": domain_type,
        "inventoryGroupId": inventory_group_id,
        **{
            stat_key: int(stats.get(stat_key) or 0)
            for stat_key in _empty_admin_data_resource_stats()
        },
    }


def _serialize_admin_data_resource(
    resource: DataResource, request_user
) -> dict[str, Any]:
    layers = list(resource.map_layers.all())
    layer = layers[0] if layers else None
    uploader = _serialize_uploader(resource.maintainer, request_user)
    maintainer = uploader["displayName"] if uploader else ""
    return {
        "id": resource.id,
        "name": resource.name,
        "code": resource.code,
        "dataType": resource.data_type,
        "domainType": resource.domain_type or None,
        "category": _serialize_dictionary_item(resource.category),
        "source": resource.source,
        "provider": resource.provider,
        "dataDate": resource.data_date.isoformat() if resource.data_date else None,
        "spatialExtent": resource.spatial_extent,
        "coordinateSystem": resource.coordinate_system,
        "fileFormat": resource.file_format,
        "storagePath": resource.storage_path,
        "description": resource.description,
        "qualityNote": resource.quality_note,
        "defaultVisualization": resource.default_visualization,
        "sizeBytes": resource.size_bytes,
        "itemCount": resource.item_count,
        "status": resource.status,
        "inventoryGroupId": resource.inventory_group_id,
        "accessGroups": [
            _serialize_access_group(group)
            for group in selectable_access_groups_for(
                resource.access_groups.all(), request_user
            )
        ],
        "canManageAccess": bool(
            has_feature_perm(request_user, "catalog.change_dataresource")
            or resource.maintainer_id == getattr(request_user, "id", None)
        ),
        "maintainer": maintainer,
        "uploader": uploader,
        "createdAt": timezone.localtime(resource.created_at).isoformat(),
        "updatedAt": timezone.localtime(resource.updated_at).isoformat(),
        "defaultLayer": _serialize_admin_resource_layer(layer),
    }


def _serialize_data_resource_group(group: DataResourceGroup) -> dict[str, Any]:
    return {
        "id": group.id,
        "name": group.name,
        "createdAt": timezone.localtime(group.created_at).isoformat(),
        "updatedAt": timezone.localtime(group.updated_at).isoformat(),
    }


def _clean_resource_group_name(value: Any) -> str | JsonResponse:
    name = str(value or "").strip()
    if not name:
        return JsonResponse({"detail": "组别名称不能为空"}, status=400)
    if len(name) > 120:
        return JsonResponse({"detail": "组别名称不能超过 120 个字符"}, status=400)
    return name


def _serialize_access_group(group: Group) -> dict[str, Any]:
    return {
        "id": group.id,
        "name": group.name,
        "isGuest": is_guest_group(group),
        "isSuperadmin": is_superadmin_group(group),
    }


def _serialize_uploader(user, viewer) -> dict[str, Any] | None:
    if user is None:
        return None
    if not user_is_visible_to(viewer, user):
        return None
    return {
        "id": user.id,
        "username": user.get_username(),
        "displayName": user.get_full_name() or user.get_username(),
    }


def _serialize_dictionary_item(item) -> dict[str, Any] | None:
    if item is None:
        return None
    return {
        "id": item.id,
        "type": item.dict_type,
        "code": item.code,
        "name": item.name,
    }


def _serialize_admin_resource_layer(layer: MapLayer | None) -> dict[str, Any] | None:
    if layer is None:
        return None
    return {
        "id": layer.id,
        "name": layer.name,
        "code": layer.code,
        "layerType": layer.layer_type,
        "geometryType": layer.geometry_type,
        "defaultVisible": layer.default_visible,
        "defaultOpacity": layer.default_opacity,
        "symbolization": layer.symbolization,
        "rasterRules": layer.raster_rules,
        "isActive": layer.is_active,
    }


def _sync_default_visualization_layer(
    resource: DataResource, visualization: dict[str, Any]
) -> JsonResponse | None:
    if resource.data_type not in {
        DataResource.DataType.VECTOR,
        DataResource.DataType.RASTER,
    }:
        return None

    layer = resource.map_layers.order_by("id").first()
    if layer is None:
        layer = MapLayer(
            code=_unique_layer_code(resource.code),
            name=resource.name,
            data_resource=resource,
            layer_type=(
                MapLayer.LayerType.RASTER
                if resource.data_type == DataResource.DataType.RASTER
                else MapLayer.LayerType.VECTOR
            ),
            source_path=resource.storage_path,
            is_active=resource.status == DataResource.Status.ACTIVE,
        )

    layer.name = str(visualization.get("layerName") or resource.name).strip()
    if not layer.name:
        return JsonResponse({"detail": "图层名称不能为空"}, status=400)
    layer.default_visible = bool(visualization.get("defaultVisible", False))
    layer.default_opacity = _default_opacity(visualization.get("defaultOpacity", 85))
    if isinstance(layer.default_opacity, JsonResponse):
        return layer.default_opacity
    layer.source_path = resource.storage_path
    layer.is_active = resource.status == DataResource.Status.ACTIVE

    symbolization = visualization.get("symbolization")
    if symbolization is not None:
        if not isinstance(symbolization, dict):
            return JsonResponse(
                {"detail": "symbolization 必须是 JSON 对象"}, status=400
            )
        layer.symbolization = symbolization

    raster_rules = visualization.get("rasterRules")
    if raster_rules is not None:
        if not isinstance(raster_rules, dict):
            return JsonResponse({"detail": "rasterRules 必须是 JSON 对象"}, status=400)
        layer.raster_rules = raster_rules
    layer.save()
    layer.access_groups.set(resource.access_groups.all())
    return None


def _default_opacity(value: Any) -> int | JsonResponse:
    try:
        opacity = int(value)
    except (TypeError, ValueError):
        return JsonResponse(
            {"detail": "defaultOpacity 必须是 0 到 100 的整数"}, status=400
        )
    if opacity < 0 or opacity > 100:
        return JsonResponse(
            {"detail": "defaultOpacity 必须是 0 到 100 的整数"}, status=400
        )
    return opacity


def _unique_layer_code(base_code: str) -> str:
    code = base_code
    if not MapLayer.objects.filter(code=code).exists():
        return code
    index = 2
    while MapLayer.objects.filter(code=f"{base_code}-layer-{index}").exists():
        index += 1
    return f"{base_code}-layer-{index}"


def _normalize_group_ids(value: Any) -> set[int] | JsonResponse:
    if not isinstance(value, list):
        return JsonResponse({"detail": "accessGroupIds 必须是数组"}, status=400)
    try:
        return {int(group_id) for group_id in value}
    except (TypeError, ValueError):
        return JsonResponse({"detail": "accessGroupIds 必须是整数数组"}, status=400)


def _delete_admin_data_resource(
    request, resource: DataResource, payload: dict[str, Any]
):
    confirmation = str(payload.get("confirmationName", "")).strip()
    if confirmation != resource.name:
        return JsonResponse({"detail": "删除确认名称与数据资源名称不一致"}, status=400)

    name = resource.name
    target_id = resource.id
    target_code = resource.code
    storage_message = _delete_imported_resource_storage(resource)
    MapLayer.objects.filter(data_resource=resource).delete()
    resource.delete()
    log_operation(
        request.user,
        "数据管理",
        "删除存量数据",
        "success",
        f"{name}；{storage_message}",
        request,
        target_type="data_resource",
        target_id=target_id,
        target_code=target_code,
        target_name=name,
    )
    return JsonResponse({"detail": "数据资源已删除"})


def _delete_imported_resource_storage(resource: DataResource) -> str:
    if not resource.storage_path:
        return "未配置存储路径，仅删除资源登记"
    if resource.data_type == DataResource.DataType.VECTOR:
        messages = [_drop_geopackage_layer(resource.storage_path)]
        dataset = VectorDataset.objects.filter(resource=resource).first()
        if dataset and dataset.source_archive_path:
            source_path = vector_original_path(dataset.source_archive_path)
            if source_path.exists() and source_path.is_file():
                source_path.unlink()
                messages.append("已清理原始矢量归档")
        return "；".join(messages)
    if (
        resource.data_type == DataResource.DataType.TABLE
        and resource.file_format.upper() == "SQLITE"
    ):
        return _drop_sqlite_table(resource.storage_path)
    if (
        resource.source in {"非地理数据目录扫描", "用户导入"}
        and "/" in resource.storage_path
    ):
        return _delete_research_file(resource.storage_path)
    return "保留原始文件，仅删除资源登记"


def _drop_geopackage_layer(layer_name: str) -> str:
    try:
        layer_name = validate_vector_layer_name(layer_name)
    except StoragePathError as exc:
        return f"跳过 GeoPackage 清理：{exc}"
    path = vector_geopackage_path()
    if not path.exists():
        return "GeoPackage 文件不存在，仅删除资源登记"
    with sqlite3.connect(path) as connection:
        quoted_name = layer_name.replace('"', '""')
        connection.execute(f'DROP TABLE IF EXISTS "{quoted_name}"')
        for table_name in (
            "gpkg_contents",
            "gpkg_geometry_columns",
            "gpkg_data_columns",
            "gpkg_extensions",
        ):
            if _sqlite_table_exists(connection, table_name):
                connection.execute(
                    f"DELETE FROM {table_name} WHERE table_name = ?",
                    (layer_name,),
                )
    return "已清理 GeoPackage 图层"


def _drop_sqlite_table(table_name: str) -> str:
    path = table_data_path("data.sqlite")
    if not path.exists():
        return "SQLite 文件不存在，仅删除资源登记"
    quoted_name = str(table_name).replace('"', '""')
    with sqlite3.connect(path) as connection:
        connection.execute(f'DROP TABLE IF EXISTS "{quoted_name}"')
        if _sqlite_table_exists(connection, "data_columns"):
            connection.execute(
                "DELETE FROM data_columns WHERE table_name = ?",
                (table_name,),
            )
    return "已清理 SQLite 表"


def _delete_research_file(relative_path: str) -> str:
    try:
        path = research_path(relative_path)
    except StoragePathError as exc:
        return f"跳过文件清理：{exc}"
    if not path.exists() or not path.is_file():
        return "文件不存在，仅删除资源登记"
    path.unlink()
    return "已删除研究数据文件"


def _admin_resource_action_label(action: str, payload: dict[str, Any]) -> str:
    if action == "setStatus":
        return "切换数据状态"
    if action == "saveVisualization":
        return "保存默认可视化方案"
    if action == "updateAccess":
        return "配置数据访问权限"
    if action == "updateInventoryGroup":
        return "配置数据组别"
    changed = []
    if "status" in payload:
        changed.append("状态")
    if "visualization" in payload:
        changed.append("默认可视化")
    if "accessGroupIds" in payload:
        changed.append("访问权限")
    if "inventoryGroupId" in payload:
        changed.append("组别")
    return f"更新存量数据{'、'.join(changed)}" if changed else "更新存量数据"


def _data_resources_csv_response(
    resources: list[DataResource], request_user
) -> HttpResponse:
    output = io.StringIO()
    output.write("\ufeff")
    headers, body = _data_resource_export_rows(resources, request_user)
    output.write(",".join(headers))
    output.write("\n")
    for row in body:
        output.write(",".join(_escape_csv_cell(cell) for cell in row))
        output.write("\n")
    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="data-inventory.csv"'
    return response


def _data_resources_xlsx_response(
    resources: list[DataResource], request_user
) -> HttpResponse:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "存量数据"
    headers, body = _data_resource_export_rows(resources, request_user)
    sheet.append(headers)
    for row in body:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="D9EDE6")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="173F39")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in sheet.columns:
        values = [
            "" if cell.value is None else str(cell.value) for cell in column_cells
        ]
        width = min(max(max((len(value) for value in values), default=8) + 2, 10), 36)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="data-inventory.xlsx"'
    return response


def _data_resource_export_rows(
    resources: list[DataResource], request_user
) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "存量数据组别",
        "数据名称",
        "数据编号",
        "数据类型",
        "业务数据类型",
        "状态",
        "分类",
        "来源",
        "提供单位",
        "上传用户",
        "上传账号",
        "数据日期",
        "空间范围",
        "坐标信息",
        "文件格式",
        "存储路径",
        "数据大小(B)",
        "数据大小",
        "数据条目数",
        "访问角色",
        "默认图层",
        "图层类型",
        "几何类型",
        "资源说明",
        "质量说明",
        "创建时间",
        "更新时间",
    ]
    body = []
    for resource in resources:
        layers = list(resource.map_layers.all())
        layer = layers[0] if layers else None
        uploader = _serialize_uploader(resource.maintainer, request_user)
        body.append(
            [
                resource.inventory_group.name
                if resource.inventory_group
                else "默认分组",
                resource.name,
                resource.code,
                resource.get_data_type_display(),
                resource.get_domain_type_display()
                if resource.domain_type
                else "未归类",
                resource.get_status_display(),
                resource.category.name if resource.category else "未分类",
                _export_text(resource.source, "未记录来源"),
                _export_text(resource.provider, "未记录提供单位"),
                uploader["displayName"] if uploader else "未记录或不可见",
                uploader["username"] if uploader else "",
                resource.data_date.isoformat() if resource.data_date else "未记录",
                _export_text(resource.spatial_extent, "未记录"),
                _export_text(resource.coordinate_system, "未记录"),
                _export_text(resource.file_format, "未记录"),
                _export_text(resource.storage_path, "未记录"),
                int(resource.size_bytes or 0),
                _format_bytes(resource.size_bytes),
                int(resource.item_count or 0),
                "、".join(
                    group.name
                    for group in visible_groups_for(
                        resource.access_groups.all(), request_user
                    )
                )
                or "未单独配置",
                layer.name if layer else "未配置",
                layer.get_layer_type_display() if layer else "",
                layer.get_geometry_type_display() if layer else "",
                _export_text(resource.description, "未记录"),
                _export_text(resource.quality_note, "未记录"),
                timezone.localtime(resource.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                timezone.localtime(resource.updated_at).strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )
    return headers, body


def _export_text(value: Any, fallback: str = "") -> str:
    text = str(value or "").strip()
    return text if text else fallback


def _format_bytes(value: Any) -> str:
    try:
        size = float(value or 0)
    except (TypeError, ValueError):
        size = 0
    units = ["B", "KB", "MB", "GB", "TB"]
    unit_index = 0
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(size)} {units[unit_index]}"
    return f"{size:.1f} {units[unit_index]}"


def _escape_csv_cell(value: Any) -> str:
    text = str(value)
    if re.search(r'[",\n]', text):
        escaped = text.replace('"', '""')
        return f'"{escaped}"'
    return text


def _sqlite_table_exists(connection: sqlite3.Connection, table_name: str) -> bool:
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def _cpu_snapshot() -> dict[str, Any]:
    logical_count = os.cpu_count() or 1
    load_average = _load_average()
    usage_percent = _cpu_usage_percent(logical_count, load_average)
    return {
        "model": _cpu_model(),
        "physicalCount": _physical_cpu_count(logical_count),
        "logicalCount": logical_count,
        "usagePercent": usage_percent,
        "loadAverage": load_average,
    }


def _cpu_model() -> str:
    if platform.system() == "Darwin":
        return _run_text(["sysctl", "-n", "machdep.cpu.brand_string"])
    if platform.system() == "Windows":
        return _run_text(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "(Get-CimInstance Win32_Processor | Select-Object -First 1).Name",
            ]
        )
    if Path("/proc/cpuinfo").exists():
        for line in Path("/proc/cpuinfo").read_text(errors="ignore").splitlines():
            if line.lower().startswith("model name"):
                return line.split(":", 1)[1].strip()
    return platform.processor() or platform.machine()


def _physical_cpu_count(logical_count: int) -> int:
    if platform.system() == "Darwin":
        value = _run_text(["sysctl", "-n", "hw.physicalcpu"])
        return _safe_int(value, logical_count)
    if platform.system() == "Windows":
        value = _run_text(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "(Get-CimInstance Win32_Processor | Measure-Object -Property NumberOfCores -Sum).Sum",
            ]
        )
        return _safe_int(value, logical_count)
    return logical_count


def _cpu_usage_percent(logical_count: int, load_average: list[float]) -> float:
    if platform.system() == "Windows":
        value = _run_text(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "(Get-CimInstance Win32_Processor | Measure-Object -Property LoadPercentage -Average).Average",
            ]
        )
        return round(float(_safe_number(value, 0)), 1)
    return round(min((load_average[0] / logical_count) * 100, 100), 1)


def _load_average() -> list[float]:
    try:
        return [round(value, 2) for value in os.getloadavg()]
    except (AttributeError, OSError):
        return [0.0, 0.0, 0.0]


def _memory_snapshot() -> dict[str, Any]:
    total = _memory_total_bytes()
    available = _memory_available_bytes()
    used = max(total - available, 0) if total else 0
    usage_percent = round((used / total) * 100, 1) if total else 0
    return {
        "model": "系统内存",
        "slotCount": 1 if total else 0,
        "totalBytes": total,
        "usedBytes": used,
        "availableBytes": available,
        "usagePercent": usage_percent,
    }


def _memory_total_bytes() -> int:
    if platform.system() == "Darwin":
        return _safe_int(_run_text(["sysctl", "-n", "hw.memsize"]), 0)
    if platform.system() == "Windows":
        return _safe_int(
            _run_text(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    "(Get-CimInstance Win32_ComputerSystem).TotalPhysicalMemory",
                ]
            ),
            0,
        )
    meminfo = _linux_meminfo()
    return meminfo.get("MemTotal", 0) * 1024


def _memory_available_bytes() -> int:
    if platform.system() == "Windows":
        value = _run_text(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "(Get-CimInstance Win32_OperatingSystem).FreePhysicalMemory",
            ]
        )
        return _safe_int(value, 0) * 1024
    if platform.system() == "Darwin":
        page_size = _safe_int(_run_text(["sysctl", "-n", "hw.pagesize"]), 4096)
        stats = _run_text(["vm_stat"])
        free_pages = 0
        for key in ("Pages free", "Pages inactive", "Pages speculative"):
            match = re.search(rf"{re.escape(key)}:\s+(\d+)\.", stats)
            if match:
                free_pages += int(match.group(1))
        return free_pages * page_size
    meminfo = _linux_meminfo()
    return meminfo.get("MemAvailable", 0) * 1024


def _linux_meminfo() -> dict[str, int]:
    meminfo_path = Path("/proc/meminfo")
    if not meminfo_path.exists():
        return {}
    values = {}
    for line in meminfo_path.read_text(errors="ignore").splitlines():
        key, _, value = line.partition(":")
        number = value.strip().split(" ")[0]
        values[key] = _safe_int(number, 0)
    return values


def _disk_snapshot() -> dict[str, Any]:
    root_usage = shutil.disk_usage(settings.BASE_DIR)
    devices = _disk_devices()
    used = root_usage.used
    total = root_usage.total
    return {
        "count": len(devices) or 1,
        "devices": devices,
        "mount": str(settings.BASE_DIR),
        "totalBytes": total,
        "usedBytes": used,
        "freeBytes": root_usage.free,
        "usagePercent": round((used / total) * 100, 1) if total else 0,
    }


def _disk_devices() -> list[dict[str, str]]:
    if platform.system() == "Darwin":
        return _darwin_disk_devices()
    if platform.system() == "Linux":
        return _linux_disk_devices()
    if platform.system() == "Windows":
        return _windows_disk_devices()
    return []


def _darwin_disk_devices() -> list[dict[str, str]]:
    text = _run_text(["diskutil", "list", "physical"])
    devices = []
    for line in text.splitlines():
        match = re.match(r"^/dev/(\S+)", line.strip())
        if match:
            device = match.group(1)
            info = _run_text(["diskutil", "info", device])
            model = ""
            size = ""
            for info_line in info.splitlines():
                if "Device / Media Name:" in info_line:
                    model = info_line.split(":", 1)[1].strip()
                if "Disk Size:" in info_line:
                    size = info_line.split(":", 1)[1].strip()
            devices.append({"name": device, "model": model, "size": size})
    return devices


def _linux_disk_devices() -> list[dict[str, str]]:
    text = _run_text(["lsblk", "-dn", "-o", "NAME,MODEL,SIZE,TYPE"])
    devices = []
    for line in text.splitlines():
        parts = line.split()
        if len(parts) >= 3 and parts[-1] == "disk":
            devices.append(
                {
                    "name": parts[0],
                    "model": " ".join(parts[1:-2]),
                    "size": parts[-2],
                }
            )
    return devices


def _windows_disk_devices() -> list[dict[str, str]]:
    text = _run_text(
        [
            "powershell",
            "-NoProfile",
            "-Command",
            'Get-CimInstance Win32_DiskDrive | ForEach-Object { "$($_.DeviceID)|$($_.Model)|$($_.Size)" }',
        ]
    )
    devices = []
    for line in text.splitlines():
        name, model, size = _split_fixed(line, "|", 3)
        if name:
            devices.append(
                {
                    "name": name,
                    "model": model,
                    "size": _format_bytes_text(_safe_int(size, 0)),
                }
            )
    return devices


def _run_text(command: list[str]) -> str:
    try:
        return subprocess.check_output(
            command,
            stderr=subprocess.DEVNULL,
            text=True,
            timeout=2,
        ).strip()
    except (OSError, subprocess.SubprocessError):
        return ""


def _safe_int(value: Any, default: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _safe_number(value: Any, default: float) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def _split_fixed(value: str, separator: str, count: int) -> list[str]:
    parts = value.split(separator)
    return [*(parts[:count]), *([""] * count)][:count]


def _format_bytes_text(value: int) -> str:
    if value <= 0:
        return ""
    units = ["B", "KB", "MB", "GB", "TB"]
    current = float(value)
    unit_index = 0
    while current >= 1024 and unit_index < len(units) - 1:
        current /= 1024
        unit_index += 1
    return f"{current:.1f} {units[unit_index]}"


def _serialize_profile(user) -> dict[str, Any]:
    profile = _ensure_profile(user)
    granted = granted_feature_permissions(user)
    disabled = disabled_feature_permissions(user)
    effective = effective_feature_permissions(user)

    # 构建头像URL
    avatar_url = profile.avatar_url
    if profile.avatar_data:
        avatar_url = f"/api/users/{user.id}/avatar/"

    return {
        "user": serialize_user(user),
        "avatarUrl": avatar_url,
        "department": profile.department,
        "grantedPermissions": sorted(granted),
        "disabledPermissions": sorted(disabled),
        "effectivePermissions": sorted(effective),
        "availablePermissions": [
            _serialize_permission(item) for item in FEATURE_PERMISSIONS
        ],
    }


def _serialize_admin_user(user, request_user=None) -> dict[str, Any]:
    serialized = serialize_user(user)
    serialized["groupIds"] = list(user.groups.values_list("id", flat=True))
    serialized["isActive"] = user.is_active
    serialized["groupPermissions"] = sorted(_group_feature_permissions(user))
    serialized["directPermissions"] = sorted(direct_feature_permissions(user))
    serialized["disabledPermissions"] = sorted(disabled_feature_permissions(user))
    serialized["effectivePermissions"] = sorted(effective_feature_permissions(user))
    serialized["operationLogGroupIds"] = _operation_log_group_ids(
        user, request_user or user
    )
    return serialized


def _group_feature_permissions(user) -> set[str]:
    permission_names = set(FEATURE_PERMISSION_NAMES)
    return {
        permission_name
        for group in user.groups.prefetch_related("permissions__content_type").all()
        for permission in group.permissions.all()
        for permission_name in [
            f"{permission.content_type.app_label}.{permission.codename}"
        ]
        if permission_name in permission_names
    }


def _serialize_group(group: Group, viewer) -> dict[str, Any]:
    permissions = {
        f"{permission.content_type.app_label}.{permission.codename}"
        for permission in group.permissions.select_related("content_type").all()
    }
    is_superadmin = is_superadmin_group(group)
    return {
        "id": group.id,
        "name": group.name,
        "userCount": visible_users_for(group.user_set.all(), viewer).count(),
        "permissions": sorted(permissions & set(FEATURE_PERMISSION_NAMES)),
        "isProtected": is_builtin_group(group),
        "lockedPermissions": sorted(
            superadmin_group_locked_permissions() if is_superadmin else set()
        ),
    }


def _available_permissions() -> list[dict[str, str]]:
    return [_serialize_permission(item) for item in FEATURE_PERMISSIONS]


def _serialize_permission(permission) -> dict[str, str]:
    return {
        "id": permission.perm_name,
        "label": permission.name,
        "group": permission.group,
    }


def _serialize_application_settings(user) -> dict[str, Any]:
    raw = load_runtime_config_document(settings.PROJECT_CONFIG)
    application = raw["application"]
    return {
        "systemName": application["system"]["name"],
        "allowRegistration": application["system"]["allow_registration"],
        "map": {
            "defaultCenter": application["map"]["default_center"],
            "defaultZoom": application["map"]["default_zoom"],
            "defaultBasemap": application["map"]["default_basemap"],
            "mapboxAccessToken": application["map"].get("mapbox_access_token", ""),
        },
        "limits": {
            "uploadMaxMb": application["limits"]["upload_max_mb"],
            "queryResultLimit": application["limits"]["query_result_limit"],
            "maxRasterSidePixels": application["limits"]["max_raster_side_pixels"],
        },
        "raster": {
            "symbolizerTimeoutSeconds": application["raster"][
                "symbolizer_timeout_seconds"
            ],
        },
        "editable": has_feature_perm(user, "core.manage_system_settings"),
    }


def _serialize_operation_log(log: OperationLog) -> dict[str, Any]:
    operator = "系统"
    if log.user_id and log.user:
        operator = log.user.get_full_name() or log.user.get_username()
    return {
        "id": log.id,
        "occurredAt": timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
        "operator": operator,
        "module": log.module,
        "action": log.action,
        "result": log.status,
        "targetType": log.target_type,
        "targetId": log.target_id,
        "targetCode": log.target_code,
        "targetName": log.target_name,
        "ipAddress": log.ip_address or "",
        "summary": log.message,
    }


def _serialize_system_log_file(path: Path, root: Path) -> dict[str, Any]:
    stat = path.stat()
    modified_at = datetime.fromtimestamp(
        stat.st_mtime, tz=timezone.get_current_timezone()
    )
    return {
        "name": path.relative_to(root).as_posix(),
        "sizeBytes": stat.st_size,
        "modifiedAt": timezone.localtime(modified_at).isoformat(),
    }


def _available_system_log_files() -> list[dict[str, Any]]:
    root = app_path("logs")
    if not root.exists():
        return []
    files = [
        path
        for path in root.iterdir()
        if path.is_file() and (path.suffix == ".log" or ".log." in path.name)
    ]
    files.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return [_serialize_system_log_file(path, root) for path in files[:50]]


def _tail_system_log_file(file_name: str, lines: int) -> str:
    safe_name = Path(file_name)
    if safe_name.is_absolute() or len(safe_name.parts) != 1 or ".." in safe_name.parts:
        raise StoragePathError("非法日志文件名")
    path = app_path("logs", safe_name.name)
    if not path.is_file():
        return ""
    max_bytes = 512 * 1024
    with path.open("rb") as handle:
        handle.seek(0, os.SEEK_END)
        size = handle.tell()
        handle.seek(max(size - max_bytes, 0), os.SEEK_SET)
        data = handle.read()
    text = data.decode("utf-8", errors="replace")
    return "\n".join(text.splitlines()[-lines:])


def _filter_backup_runs(queryset, params):
    plan_type = str(params.get("planType", "")).strip()
    target_type = str(params.get("targetType", "")).strip()
    status = str(params.get("status", "")).strip()
    if plan_type:
        if plan_type not in BackupRun.PlanType.values:
            return JsonResponse(
                {"detail": "planType 仅支持 platform 或 research"}, status=400
            )
        queryset = queryset.filter(plan_type=plan_type)
    if target_type:
        if target_type not in BackupRun.TargetType.values:
            return JsonResponse(
                {"detail": "targetType 仅支持 local 或 object_storage"}, status=400
            )
        queryset = queryset.filter(target_type=target_type)
    if status:
        if status not in BackupRun.Status.values:
            return JsonResponse(
                {"detail": "status 仅支持 queued、running、success 或 failed"},
                status=400,
            )
        queryset = queryset.filter(status=status)
    return queryset


def _filter_operation_logs(queryset, params):
    user_id = params.get("userId")
    operator = str(params.get("operator", "")).strip()
    module = str(params.get("module", "")).strip()
    action = str(params.get("action", "")).strip()
    result = str(params.get("result", "")).strip()
    target_type = str(params.get("targetType", "")).strip()
    target_id = str(params.get("targetId", "")).strip()
    keyword = str(params.get("keyword", "")).strip()
    start_time = _parse_query_datetime(params.get("startTime"), end_of_day=False)
    end_time = _parse_query_datetime(params.get("endTime"), end_of_day=True)

    if user_id not in (None, ""):
        try:
            queryset = queryset.filter(user_id=int(user_id))
        except (TypeError, ValueError):
            return queryset.none()
    if operator:
        queryset = queryset.filter(
            Q(user__username__icontains=operator)
            | Q(user__first_name__icontains=operator)
            | Q(user__last_name__icontains=operator)
        )
    if module:
        queryset = queryset.filter(module__icontains=module)
    if action:
        queryset = queryset.filter(action__icontains=action)
    if result:
        queryset = queryset.filter(status=result)
    if target_type:
        queryset = queryset.filter(target_type=target_type)
    if target_id:
        try:
            queryset = queryset.filter(target_id=int(target_id))
        except ValueError:
            return queryset.none()
    if keyword:
        queryset = queryset.filter(
            Q(user__username__icontains=keyword)
            | Q(user__first_name__icontains=keyword)
            | Q(user__last_name__icontains=keyword)
            | Q(module__icontains=keyword)
            | Q(action__icontains=keyword)
            | Q(target_type__icontains=keyword)
            | Q(target_code__icontains=keyword)
            | Q(target_name__icontains=keyword)
            | Q(message__icontains=keyword)
        )
    if start_time:
        queryset = queryset.filter(created_at__gte=start_time)
    if end_time:
        queryset = queryset.filter(created_at__lte=end_time)
    return queryset


def _scope_operation_logs(queryset, user):
    if has_feature_perm(user, "core.view_all_operation_logs"):
        return visible_operation_logs_for(queryset, user)

    scope = Q(user_id=user.id)
    if has_feature_perm(user, "core.view_group_operation_logs"):
        group_ids = _operation_log_group_ids(user)
        if group_ids:
            scope |= Q(user__groups__id__in=group_ids)
    return visible_operation_logs_for(queryset.filter(scope), user).distinct()


def _operation_log_group_ids(user, viewer=None) -> list[int]:
    try:
        profile = user.profile
    except ObjectDoesNotExist:
        return []
    group_ids = profile.operation_log_group_ids
    if not isinstance(group_ids, list):
        return []
    normalized = []
    for group_id in group_ids:
        try:
            normalized.append(int(group_id))
        except (TypeError, ValueError):
            continue
    return visible_group_ids_for(normalized, viewer or user)


def _operation_log_group_ids_payload(value: Any, viewer) -> list[int] | JsonResponse:
    if not isinstance(value, list):
        return JsonResponse({"detail": "operationLogGroupIds 必须是数组"}, status=400)
    try:
        normalized_group_ids = sorted({int(group_id) for group_id in value})
    except (TypeError, ValueError):
        return JsonResponse(
            {"detail": "operationLogGroupIds 必须是整数数组"}, status=400
        )
    if not normalized_group_ids:
        return []
    existing_ids = set(
        visible_groups_for(
            Group.objects.filter(id__in=normalized_group_ids), viewer
        ).values_list("id", flat=True)
    )
    missing_ids = set(normalized_group_ids) - existing_ids
    if missing_ids:
        return JsonResponse({"detail": "包含不存在的日志角色"}, status=400)
    return normalized_group_ids


def _set_operation_log_group_ids(user, group_ids: list[int]) -> None:
    profile = _ensure_profile(user)
    profile.operation_log_group_ids = group_ids
    profile.save(update_fields=["operation_log_group_ids", "updated_at"])


def _set_disabled_permissions(user, permission_names: list[str]) -> None:
    granted = granted_feature_permissions(user)
    profile = _ensure_profile(user)
    profile.disabled_permissions = sorted(
        permission_name
        for permission_name in permission_names
        if permission_name in granted
    )
    profile.save(update_fields=["disabled_permissions", "updated_at"])


def _parse_query_datetime(value: Any, *, end_of_day: bool):
    if not value:
        return None
    raw_value = str(value)
    parsed = parse_datetime(raw_value)
    if parsed is None:
        parsed_date = parse_date(raw_value)
        if parsed_date is None:
            return None
        time_value = time.max if end_of_day else time.min
        parsed = datetime.combine(parsed_date, time_value)
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _positive_query_int(value: Any, *, default: int) -> int | JsonResponse:
    if value in (None, ""):
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return JsonResponse({"detail": "分页参数必须是正整数"}, status=400)
    if parsed <= 0:
        return JsonResponse({"detail": "分页参数必须是正整数"}, status=400)
    return parsed


def _create_admin_user(User, payload: dict[str, Any], request_user=None):
    username = _required_string(payload.get("username"), "username")
    if isinstance(username, JsonResponse):
        return username

    # 自动生成密码
    password = generate_password()

    display_name = str(payload.get("displayName", "")).strip()
    department = str(payload.get("department", "")).strip()
    is_active = bool(payload.get("isActive", True))
    group_ids = payload.get("groupIds", [])
    if not isinstance(group_ids, list):
        return JsonResponse({"detail": "groupIds 必须是数组"}, status=400)
    try:
        normalized_group_ids = {int(group_id) for group_id in group_ids}
    except (TypeError, ValueError):
        return JsonResponse({"detail": "groupIds 必须是整数数组"}, status=400)
    if not normalized_group_ids:
        return JsonResponse({"detail": "角色为必选项"}, status=400)

    groups = list(
        visible_groups_for(
            Group.objects.filter(id__in=normalized_group_ids),
            request_user,
        )
    )
    if len(groups) != len(normalized_group_ids):
        return JsonResponse({"detail": "包含不存在的角色"}, status=400)
    if any(is_superadmin_group(group) for group in groups):
        return JsonResponse(
            {
                "detail": f"不能将{DEFAULT_USER_GROUP_NAME}加入{SUPERADMIN_GROUP_NAME}角色"
            },
            status=400,
        )
    if (
        request_user is not None
        and not is_superadmin_user(request_user)
        and any(is_platform_admin_group(group) for group in groups)
    ):
        return JsonResponse(
            {"detail": "只有超级管理员可以分配平台管理员角色"}, status=400
        )

    try:
        email = validate_account_email(payload.get("email"))
    except AccountEmailError as exc:
        return JsonResponse({"detail": str(exc)}, status=400)

    try:
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=display_name,
                is_active=is_active,
            )
            user.groups.set(groups)
            profile = _ensure_profile(user)
            profile.department = department
            profile.normalized_email = email
            profile.save(update_fields=["department", "normalized_email", "updated_at"])
    except IntegrityError:
        if User.objects.filter(email__iexact=email).exists():
            return JsonResponse({"detail": "邮箱已被使用"}, status=400)
        return JsonResponse({"detail": "用户名已存在"}, status=400)
    return user, password


def _groups():
    return Group.objects.prefetch_related("permissions", "user_set").order_by("name")


def _ensure_profile(user):
    try:
        return user.profile
    except ObjectDoesNotExist:
        return UserProfile.objects.create(user=user)


def _backup_superadmin_denied(request) -> JsonResponse | None:
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "请先登录"}, status=401)
    if not has_feature_perm(
        request.user, "core.manage_data_backup"
    ) or not is_superadmin_user(request.user):
        return JsonResponse({"detail": "数据备份属于系统级维护功能"}, status=403)
    return None


def _json_payload(request) -> dict[str, Any] | JsonResponse:
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "请求体不是有效 JSON"}, status=400)
    if not isinstance(payload, dict):
        return JsonResponse({"detail": "请求体必须是 JSON 对象"}, status=400)
    return payload


def _required_string(value: Any, key: str) -> str | JsonResponse:
    if not isinstance(value, str) or not value.strip():
        return JsonResponse({"detail": f"{key} 必须是非空字符串"}, status=400)
    return value.strip()


def _permission_names(value: Any) -> list[str] | JsonResponse:
    if not isinstance(value, list):
        return JsonResponse({"detail": "permissions 必须是数组"}, status=400)
    names = [str(item) for item in value]
    invalid = sorted(set(names) - set(FEATURE_PERMISSION_NAMES))
    if invalid:
        return JsonResponse({"detail": f"权限不存在：{', '.join(invalid)}"}, status=400)
    return names


def _set_group_feature_permissions(group: Group, permission_names: list[str]) -> None:
    permissions = feature_permission_queryset()
    feature_ids = set(permissions.values_list("id", flat=True))
    permissions_by_name = {
        f"{permission.content_type.app_label}.{permission.codename}": permission.id
        for permission in permissions
    }
    selected_ids = {
        permissions_by_name[permission]
        for permission in permission_names
        if permission in permissions_by_name
    }
    non_feature_ids = set(
        group.permissions.exclude(id__in=feature_ids).values_list("id", flat=True)
    )
    group.permissions.set([*non_feature_ids, *selected_ids])


def _set_user_feature_permissions(user, permission_names: list[str]) -> None:
    permissions = feature_permission_queryset()
    feature_ids = set(permissions.values_list("id", flat=True))
    permissions_by_name = {
        f"{permission.content_type.app_label}.{permission.codename}": permission.id
        for permission in permissions
    }
    selected_ids = {
        permissions_by_name[permission]
        for permission in permission_names
        if permission in permissions_by_name
    }
    non_feature_ids = set(
        user.user_permissions.exclude(id__in=feature_ids).values_list("id", flat=True)
    )
    user.user_permissions.set([*non_feature_ids, *selected_ids])


def _map_patch(value: Any) -> dict[str, Any] | JsonResponse:
    if not isinstance(value, dict):
        return JsonResponse({"detail": "map 必须是对象"}, status=400)
    patch: dict[str, Any] = {}
    if "defaultCenter" in value:
        center = value["defaultCenter"]
        if not isinstance(center, list | tuple) or len(center) != 2:
            return JsonResponse(
                {"detail": "defaultCenter 必须是 [经度, 纬度]"}, status=400
            )
        longitude = _finite_float(center[0], "defaultCenter[0]")
        if isinstance(longitude, JsonResponse):
            return longitude
        latitude = _finite_float(center[1], "defaultCenter[1]")
        if isinstance(latitude, JsonResponse):
            return latitude
        if longitude < -180 or longitude > 180:
            return JsonResponse({"detail": "defaultCenter 经度超出范围"}, status=400)
        if latitude < -90 or latitude > 90:
            return JsonResponse({"detail": "defaultCenter 纬度超出范围"}, status=400)
        patch["default_center"] = [longitude, latitude]
    if "defaultZoom" in value:
        default_zoom = _finite_float(value["defaultZoom"], "defaultZoom")
        if isinstance(default_zoom, JsonResponse):
            return default_zoom
        patch["default_zoom"] = default_zoom
    if "defaultBasemap" in value:
        default_basemap = _required_string(value["defaultBasemap"], "defaultBasemap")
        if isinstance(default_basemap, JsonResponse):
            return default_basemap
        patch["default_basemap"] = default_basemap
    if "mapboxAccessToken" in value:
        patch["mapbox_access_token"] = str(value["mapboxAccessToken"]).strip()
    return patch


def _limits_patch(value: Any) -> dict[str, Any] | JsonResponse:
    if not isinstance(value, dict):
        return JsonResponse({"detail": "limits 必须是对象"}, status=400)
    patch: dict[str, Any] = {}
    if "uploadMaxMb" in value:
        upload_max_mb = _positive_int(value["uploadMaxMb"], "uploadMaxMb")
        if isinstance(upload_max_mb, JsonResponse):
            return upload_max_mb
        patch["upload_max_mb"] = upload_max_mb
    if "queryResultLimit" in value:
        query_result_limit = _positive_int(
            value["queryResultLimit"],
            "queryResultLimit",
        )
        if isinstance(query_result_limit, JsonResponse):
            return query_result_limit
        patch["query_result_limit"] = query_result_limit
    if "maxRasterSidePixels" in value:
        max_raster_side_pixels = _positive_int(
            value["maxRasterSidePixels"],
            "maxRasterSidePixels",
        )
        if isinstance(max_raster_side_pixels, JsonResponse):
            return max_raster_side_pixels
        patch["max_raster_side_pixels"] = max_raster_side_pixels
    return patch


def _raster_patch(value: Any) -> dict[str, Any] | JsonResponse:
    if not isinstance(value, dict):
        return JsonResponse({"detail": "raster 必须是对象"}, status=400)
    patch: dict[str, Any] = {}
    if "symbolizerTimeoutSeconds" in value:
        symbolizer_timeout_seconds = _positive_int(
            value["symbolizerTimeoutSeconds"],
            "symbolizerTimeoutSeconds",
        )
        if isinstance(symbolizer_timeout_seconds, JsonResponse):
            return symbolizer_timeout_seconds
        patch["symbolizer_timeout_seconds"] = symbolizer_timeout_seconds
    return patch


def _positive_int(value: Any, key: str) -> int | JsonResponse:
    if not isinstance(value, int) or value <= 0:
        return JsonResponse({"detail": f"{key} 必须是正整数"}, status=400)
    return value


def _boolean_value(value: Any, key: str) -> bool | JsonResponse:
    if not isinstance(value, bool):
        return JsonResponse({"detail": f"{key} 必须是布尔值"}, status=400)
    return value


def _finite_float(value: Any, key: str) -> float | JsonResponse:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return JsonResponse({"detail": f"{key} 必须是有效数字"}, status=400)
    if not math.isfinite(number):
        return JsonResponse({"detail": f"{key} 必须是有效数字"}, status=400)
    return number
